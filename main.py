import datetime
import locale

import openpyxl as xl
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Border, borders
from openpyxl.utils import get_column_letter

locale.setlocale(locale.LC_ALL, 'fr_FR')

with open('input/input1.html', encoding='utf8') as file:
    html_content = file.read()

bourso = pd.DataFrame()

soup = BeautifulSoup(html_content, 'html.parser')
movements_list = soup.find('ul', class_='list__movement')
current_line = movements_list.find_next('li')
current_date = ''
while current_line is not None:
    if current_line.has_attr('class'):
        if current_line['class'][0] == 'list-operation-date-line':
            current_date = pd.to_datetime(current_line.get_text(strip=True), format='%d %B %Y')
            # print(current_date)
        elif current_line['class'][0] == 'list-operation-item':
            amount = current_line.find('div', class_="list-operation-item__amount")
            amount = float(
                amount.get_text(strip=True).replace('\xa0', '').replace('€', '').replace(',', '.').replace('−', '-'))
            description = current_line.find('span', class_='list__movement--label-user').get_text(strip=True)
            current_df = pd.DataFrame({'Date': [current_date], 'Description': [description], 'Montant': [amount]})
            bourso = pd.concat([bourso, current_df], ignore_index=True)
            # print(description)
    current_line = current_line.find_next_sibling('li')
amounts = movements_list.find_all('div', class_="list-operation-item__amount")
amounts = [float(e.get_text(strip=True).replace('\xa0', '').replace('€', '').replace(',', '.').replace('−', '-')) for e
           in amounts]

bourso.to_excel('output/output.xlsx')

manager = pd.read_excel('input/Money Manager - Excel 1-1-22 ~ 12-31-22 (1).xlsx', engine='openpyxl',
                        usecols=['Date', 'Account', 'Category', 'Subcategory', 'Note', 'EUR', 'Income/Expense'])
manager['Date'] = pd.to_datetime(manager['Date'], dayfirst=True)

manager["Montant"] = manager['EUR'] * (2 * (manager['Income/Expense'] == 'Income') - 1)
manager = manager[manager['Account'] == 'Carte']

manager = manager.drop(['EUR', 'Income/Expense'], axis=1)

manager['Pointé'] = False
bourso['Pointé'] = False

link_df = pd.DataFrame()
print(bourso)
print(manager)

link_dic = {}

for days_diff in range(30):
    print(f'Liaison des opérations à {days_diff} jours de différence...')
    for _, bourso_row in bourso[bourso['Pointé'] == False].iterrows():
        manager_filtered = manager[(manager['Pointé'] == False) &
                                   (abs(bourso_row['Date'] - manager['Date']) < datetime.timedelta(
                                       days=days_diff + 1)) &
                                   (abs(bourso_row['Date'] - manager['Date']) > datetime.timedelta(days=days_diff - 1))]
        for _, manager_row in manager_filtered.iterrows():
            if abs(bourso_row['Date'] - manager_row['Date']).days == days_diff:
                if bourso_row['Montant'] == manager_row['Montant']:
                    manager.loc[manager_row.name, 'Pointé'] = True
                    bourso.loc[bourso_row.name, 'Pointé'] = True
                    link_dic[bourso_row.name] = manager_row.name
                    break

# Construction du dataframe link
for index, value in link_dic.items():
    current_row = pd.DataFrame({
        'Date Manager': [manager.loc[value, 'Date']],
        'Account Manager': [manager.loc[value, 'Account']],
        'Category Manager': [manager.loc[value, 'Category']],
        'Subcategory Manager': [manager.loc[value, 'Subcategory']],
        'Note Manager': [manager.loc[value, 'Note']],
        'Montant Manager': [manager.loc[value, 'Montant']],
        'Montant Bourso': [bourso.loc[index, 'Montant']],
        'Date Bourso': [bourso.loc[index, 'Date']],
        'Description Bourso': [bourso.loc[index, 'Description']],
        'Jours de différence': [(bourso.loc[index, 'Date'] - manager.loc[value, 'Date']).days]
    })
    link_df = pd.concat([link_df, current_row])

for _, bourso_row in bourso[bourso['Pointé'] == False].iterrows():
    current_row = pd.DataFrame({
        'Montant Bourso': [bourso_row['Montant']],
        'Date Bourso': [bourso_row['Date']],
        'Description Bourso': [bourso_row['Description']],
    })
    link_df = pd.concat([link_df, current_row])

for _, manager_row in manager[manager['Pointé'] == False].iterrows():
    current_row = pd.DataFrame({
        'Date Manager': [manager_row['Date']],
        'Account Manager': [manager_row['Account']],
        'Category Manager': [manager_row['Category']],
        'Subcategory Manager': [manager_row['Subcategory']],
        'Note Manager': [manager_row['Note']],
        'Montant Manager': [manager_row['Montant']],
    })
    link_df = pd.concat([link_df, current_row])

link_df = link_df.iloc[link_df['Date Bourso'].fillna(link_df['Date Manager']).argsort()[::-1]]
print(link_dic)
print(bourso)
print(manager)
print(link_df)
manager.to_excel('output/manager.xlsx')
bourso.to_excel('output/bourso.xlsx')
link_df.to_excel('output/linked.xlsx')


def best_fit_column_width(sheet):
    for column_cells in sheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            sheet.column_dimensions[new_column_letter].width = new_column_length * 1.23


def apply_short_date_format(column_letter):
    for cell in ws[column_letter]:
        cell.number_format = 'DD/MM/YYYY'


wb = xl.load_workbook('output/linked.xlsx')
ws = wb.active
redfill = PatternFill(start_color='f2dcdb',
                      end_color='f2dcdb',
                      fill_type='solid')
red_border = Border(top=borders.Side(color='8d3a38', style='thin'), bottom=borders.Side(color='8d3a38', style='thin'))
ws.conditional_formatting.add('a2:K809', FormulaRule(formula=['ISBLANK($K2)'], fill=redfill, border=red_border))
apply_short_date_format('B')
apply_short_date_format('I')
best_fit_column_width(ws)
wb.save('output/linked.xlsx')
